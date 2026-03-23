import streamlit as st
import pandas as pd
import io
import re
from openpyxl import load_workbook
import datetime

# --- 1. 页面配置 ---
st.set_page_config(page_title="Cupshe 亚马逊优惠券助手", layout="wide")

# --- 2. 状态初始化 ---
if 'coupon_pool' not in st.session_state:
    st.session_state.coupon_pool = []
if 'field_configs' not in st.session_state:
    st.session_state.field_configs = []

# --- 3. 核心工具函数 ---

def clean_asin_format(raw_text):
    """【需求】无论输入什么，输出必须是英文分号分隔"""
    if not raw_text: return "", 0
    tokens = re.split(r'[;；,，\s\n\r]+', str(raw_text).strip())
    clean_list = [t.strip().upper() for t in tokens if t.strip()]
    seen = set()
    final_list = [x for x in clean_list if not (x in seen or seen.add(x))]
    return ";".join(final_list), len(final_list)

def detect_date_format(sample_val):
    """【精准修正】根据截图，支持 YYYY-MM-DD 格式识别"""
    if not sample_val:
        return "%Y-%m-%d"  # 默认匹配你截图的格式
    
    s = str(sample_val).strip()
    
    # 如果包含连字符 - (如 2021-06-20)
    if "-" in s:
        parts = s.split("-")
        if len(parts[0]) == 4: return "%Y-%m-%d" # 年-月-日
        if len(parts[2]) == 4: return "%d-%m-%Y" # 日-月-年
        
    # 如果包含斜杠 / (备用识别)
    if "/" in s:
        parts = s.split("/")
        if len(parts[0]) == 4: return "%Y/%m/%d"
        if len(parts[2]) == 4: return "%m/%d/%Y"
        
    return "%Y-%m-%d" # 兜底格式

# --- 4. 侧边栏与模板解析 ---
st.sidebar.header("🛠️ 控制中心")
mode = st.sidebar.radio("选择操作阶段", ["第一阶段：需求录入", "第二阶段：校验与导出"])

template_file = st.sidebar.file_uploader("上传 Coupon 原始模板 (Excel)", type=['xlsx'])

if template_file and not st.session_state.field_configs:
    wb = load_workbook(template_file, data_only=True)
    ws = wb.active
    configs = []
    # 扫描前25列
    for col in range(1, 26):
        title = ws.cell(row=7, column=col).value
        hint = ws.cell(row=5, column=col).value
        # 抓取第8, 9行示例作为格式参考
        sample_8 = ws.cell(row=8, column=col).value
        sample_9 = ws.cell(row=9, column=col).value
        
        if title:
            title_str = str(title).strip()
            is_drop = any(k in title_str for k in ["折扣类型", "限购", "限制", "买家", "叠加", "类型"])
            
            # 日期列判定
            date_fmt = None
            if any(k in title_str for k in ["日期", "Date"]):
                date_fmt = detect_date_format(sample_8 or sample_9)
            
            configs.append({
                "col": col, 
                "label": title_str, 
                "hint": str(hint).strip() if hint else "遵循第5行亚马逊规则",
                "is_dropdown": is_drop, 
                "options": [str(sample_8), str(sample_9)] if is_drop else [],
                "date_format": date_fmt
            })
    st.session_state.field_configs = configs

# --- 5. 第一阶段：需求录入 ---
if mode == "第一阶段：需求录入":
    if not st.session_state.field_configs:
        st.info("👋 请先在侧边栏上传模板。")
    else:
        with st.form("entry_form", clear_on_submit=True):
            st.subheader("📝 录入新需求")
            user_input_raw = {}
            grid = st.columns(2)
            
            for i, cfg in enumerate(st.session_state.field_configs):
                with grid[i % 2]:
                    label = cfg['label']
                    hint = cfg['hint']
                    if cfg['is_dropdown']:
                        user_input_raw[cfg['col']] = st.selectbox(label, options=cfg['options'], help=hint)
                    elif cfg['date_format']:
                        # 日期输入框
                        user_input_raw[cfg['col']] = st.date_input(label, help=f"自动匹配模板格式: {cfg['date_format']}")
                    elif "ASIN" in label.upper():
                        user_input_raw[cfg['col']] = st.text_area(label, help=hint, placeholder="支持直接粘贴ASIN列")
                    else:
                        user_input_raw[cfg['col']] = st.text_input(label, help=hint)
            
            if st.form_submit_button("➕ 确认添加并同步格式"):
                new_row = {}
                for c_idx, val in user_input_raw.items():
                    cfg = next(c for c in st.session_state.field_configs if c['col'] == c_idx)
                    
                    # 1. 强制日期格式输出 (YYYY-MM-DD)
                    if cfg['date_format'] and isinstance(val, (datetime.date, datetime.datetime)):
                        new_row[c_idx] = val.strftime(cfg['date_format'])
                    
                    # 2. 强制 ASIN 格式输出 (分号分隔)
                    elif "ASIN" in cfg['label'].upper():
                        final_asin, count = clean_asin_format(val)
                        new_row[c_idx] = final_asin
                        st.toast(f"已转换 {count} 个 ASIN")
                        
                    # 3. 普通文本
                    else:
                        new_row[c_idx] = str(val) if val is not None else ""
                
                st.session_state.coupon_pool.append(new_row)
                st.success("✅ 已添加！格式已自动调整。")

    # 实时预览
    if st.session_state.coupon_pool:
        st.divider()
        st.subheader("📋 预览（请确认日期格式是否为 2021-06-20 形式）")
        display_map = {c['col']: c['label'] for c in st.session_state.field_configs}
        st.dataframe(pd.DataFrame(st.session_state.coupon_pool).rename(columns=display_map))

# --- 6. 第二阶段：导出流（保持原有导出逻辑不变） ---
elif mode == "第二阶段：校验与导出":
    if not st.session_state.coupon_pool:
        st.info("需求池目前是空的。")
    else:
        if st.button("🚀 生成并下载填充好的 Excel"):
            template_file.seek(0)
            wb_out = load_workbook(template_file)
            ws_out = wb_out.active
            
            # 自动定位空行
            curr_r = 8
            while ws_out.cell(row=curr_r, column=1).value:
                curr_r += 1
            
            for row_data in st.session_state.coupon_pool:
                for col_idx, value in row_data.items():
                    ws_out.cell(row=curr_r, column=int(col_idx)).value = value
                curr_r += 1
            
            buf = io.BytesIO()
            wb_out.save(buf)
            st.download_button("💾 点击下载最终文件", buf.getvalue(), f"Coupon_Export_{datetime.date.today()}.xlsx")
